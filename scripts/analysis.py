from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sys, os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils.db_utils import create_mysql_connection
from collections import Counter


def fetch_table_data(query):
    """
    Execute the given SQL query and return column names and fetched data.
    """
    conn = create_mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query)
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    cursor.close()
    conn.close()
    return columns, data


def write_sheet(workbook, sheet_name, columns, data):
    """
    Create a worksheet with given sheet_name, write the column headers and data,
    and auto-adjust column widths based on content.
    """
    ws = workbook.create_sheet(title=sheet_name)
    ws.append(columns)

    for row in data:
        ws.append([row[col] for col in columns])

    for col_idx, _ in enumerate(columns, start=1):
        max_length = max(
            (len(str(row[columns[col_idx - 1]])) for row in data), default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4


def generate_summary_metrics():
    """
    Generate key summary metrics from the database for index performance and composition.
    Returns a list of tuples (metric_name, value).
    """
    conn = create_mysql_connection()
    cursor = conn.cursor(dictionary=True)
    metrics = []

    # Total Trading Days
    cursor.execute("SELECT COUNT(*) AS count FROM index_values")
    metrics.append(("Total Trading Days", cursor.fetchone()["count"]))

    # Total Composition Changes
    cursor.execute("""
        SELECT COUNT(*) AS total_changes 
        FROM composition_changes 
        WHERE COALESCE(added, '') <> '' OR COALESCE(removed, '') <> ''
    """)
    metrics.append(("Total Composition Changes", cursor.fetchone()["total_changes"]))

    # Total Return (%)
    cursor.execute("SELECT cumulative_return FROM index_values ORDER BY date DESC LIMIT 1")
    result = cursor.fetchone()
    total_return = round(result["cumulative_return"], 2) if result else 0.0
    metrics.append(("Total Return (%)", f"{total_return}%"))

    # Average Daily Return (%)
    cursor.execute("SELECT AVG(daily_return) AS avg_return FROM index_values WHERE daily_return IS NOT NULL")
    avg_return = cursor.fetchone()["avg_return"]
    metrics.append(("Average Daily Return (%)", f"{round(avg_return, 2)}%" if avg_return else "0%"))

    # Volatility (%)
    cursor.execute("SELECT STDDEV(daily_return) AS volatility FROM index_values WHERE daily_return IS NOT NULL")
    volatility = cursor.fetchone()["volatility"]
    metrics.append(("Volatility (%)", f"{round(volatility, 2)}%" if volatility else "0%"))

    # Best Day
    cursor.execute("SELECT date, daily_return FROM index_values WHERE daily_return IS NOT NULL ORDER BY daily_return DESC LIMIT 1")
    best = cursor.fetchone()
    if best:
        metrics.append(("Best Day", best["date"].strftime("%Y-%m-%d")))
        metrics.append(("Best Daily Return (%)", f"{round(best['daily_return'], 2)}%"))

    # Worst Day
    cursor.execute("SELECT date, daily_return FROM index_values WHERE daily_return IS NOT NULL ORDER BY daily_return ASC LIMIT 1")
    worst = cursor.fetchone()
    if worst:
        metrics.append(("Worst Day", worst["date"].strftime("%Y-%m-%d")))
        metrics.append(("Worst Daily Return (%)", f"{round(worst['daily_return'], 2)}%"))

    # Most Stable Stocks (100% presence)
    cursor.execute("SELECT COUNT(DISTINCT date) AS total_days FROM daily_composition")
    total_days = cursor.fetchone()["total_days"]
    cursor.execute("""
        SELECT ticker, COUNT(*) as appearances
        FROM daily_composition
        GROUP BY ticker
        HAVING appearances = %s
        LIMIT 5
    """, (total_days,))
    stable_stocks = ", ".join([row["ticker"] for row in cursor.fetchall()])
    metrics.append(("Most Stable Stocks (100% presence)", stable_stocks if stable_stocks else "None"))

    # Most Frequently Removed Stocks
    cursor.execute("SELECT removed FROM composition_changes WHERE removed IS NOT NULL AND removed != ''")
    removed_all = []
    for row in cursor.fetchall():
        removed_all.extend(row["removed"].split(","))
    removed_freq = Counter(removed_all)
    top_removed = ", ".join([f"{ticker}({count})" for ticker, count in removed_freq.most_common(3)])
    metrics.append(("Most Frequently Removed Stocks", top_removed if top_removed else "None"))

    # Average Number of Daily Ticker Changes
    cursor.execute("SELECT added, removed FROM composition_changes")
    total_changes = 0
    days = 0
    for row in cursor.fetchall():
        added = row["added"].split(",") if row["added"] else []
        removed = row["removed"].split(",") if row["removed"] else []
        total_changes += len(added) + len(removed)
        days += 1
    avg_changes = round(total_changes / days, 2) if days > 0 else 0
    metrics.append(("Avg # of Daily Ticker Changes", avg_changes))

    cursor.close()
    conn.close()

    return metrics


def export_all_to_excel(output_path="output/index_report.xlsx"):
    """
    Export all relevant tables and summary metrics to an Excel file.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    queries = {
        "index_performance": "SELECT * FROM index_values ORDER BY date",
        "daily_composition": "SELECT * FROM daily_composition ORDER BY date, ticker",
        "composition_changes": "SELECT * FROM composition_changes ORDER BY date",
    }

    for sheet_name, query in queries.items():
        columns, data = fetch_table_data(query)
        write_sheet(wb, sheet_name, columns, data)

    # Add summary metrics sheet
    summary_ws = wb.create_sheet(title="summary_metrics")
    for metric in generate_summary_metrics():
        summary_ws.append(metric)

    wb.save(output_path)
    print(f"✅ Excel file generated: {output_path}")


if __name__ == "__main__":
    export_all_to_excel()
